#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import socket
import requests
import sys
import getopt
import datetime
import pdb
import os
from lxml import html
from openpyxl import Workbook
from openpyxl import load_workbook
#pdb.set_trace()

'''
pip3 install requests lxml openpyxl  #安装脚本依赖程序包
调用方法：./spider.py -n <number> -u <url>  #不使用命令行选项时使用程序'默认参数'
注意：每次使用后不要删除生成的Excel(作为历史记录)，否则会重复爬取已爬取过的数据，可自行在该Excel中最后增加一列备注是否已看过方便筛选使用
'''

#默认参数
spider_num = 5  #设置(默认)每次最多爬取的数量(-1表示无限循环，若实际没这么多，会提前结束)
spider_url = 'https://www.baidu.com'  #设置(默认)需要爬取的地址(类似如：https://avmask.com/cn、https://avmask.com/cn/star/...、https://avmask.com/cn/search/...等主页、子页爬取，不支持其他站点)
spider_enable_proxy = False  #是否使用代理爬取
#在MySplider类构造函数中，'参数变量'部分修改其他参数

#全局变量
title = ['0id', '1关键信息', '2URL地址', '3占用状态']  #标题--字段意义
lists = list()  #URL队列，同样作为输出结果
key_information = list()  #存储已经爬过的所有关键信息，用来防止重复爬取

class MySplider(object):
    def __init__(self):
        #运行时变量
        self.host = ''
        self.host_complete = ''
        self.host_protocol = ''
        self.obj = ''
        self.tree = ''
        self.id = 0
        self.home_list = ['/released', '/star', '/popular', '/genre']
        self.os_name = os.name

        #参数变量
        self.port = 80  #网站端口
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9"
            }  #request头设置，不能少了Accept-Language参数，否则搜索引擎报错403
        self.proxy = {
            'http': 'http://127.0.0.1:8087',
            'https': 'http://127.0.0.1:8087'
            }  #在被墙的情况下使用http代理爬取，这里设置代理地址
        self.sheet_name = ['Sheet1']  #设置工作表名称
        self.file_name = 'data.xlsx'  #设置Excel文件名称
        self.file_path = os.path.abspath('.')  #设置Excel所在位置为程序当前所在目录
        self.file_complete_path = self.file_path + '/' + self.file_name  #设置Excel完整文件路径

        #Excel对象
        #self.wb = Workbook()  #设置工作表
        #self.initExcel()  #初始化工作表

    def initExcel(self):
        if (os.path.isfile(self.file_complete_path) == True):  #如果表格存在替换wb对象，并初始化key_information列表
            self.wb = load_workbook(self.file_complete_path)
            for cell in self.wb[self.sheet_name[0]]['B']:  #循环获得第二列：关键信息
                key_information.append(cell.value)
            self.id = self.wb[self.sheet_name[0]].cell(row = self.wb[self.sheet_name[0]].max_row, column = 1).value
            key_information.pop(0)  #删除标题行
            #print ('key_information =', key_information)
        else:
            self.wb = Workbook()
            self.wb.active.title = self.sheet_name[0]
            self.wb.active.append(title)
            self.wb.save(self.file_complete_path)

    def protocol(self, url):  #判断协议类型
        if (url.find('http://') == 0):
            return 1
        if (url.find('https://') == 0):
            return 2
        if (url.find('https://') != 0 and url.find('http://') != 0):
            return 0

    def analysis(self, url):  #解析URL，获得主机名和资源路径
        #URL的一般格式：protocol://hostname[:port]/path/[?query]#fragment
        if (len(url) <= 7 or self.protocol(url) == 0):  #url为空或数据错误返回失败
            print(self.__class__.__name__, '::', sys._getframe().f_code.co_name, '解析url错误：', url)
            return False

        #获得主机名和资源路径
        if (self.protocol(url) == 1 and url.find('/', 7) != -1):  #判断是http协议
            self.host = url[7:url.find('/', 7)]
            self.obj = url[url.find('/', 7):]
            self.host_protocol = url[:url.find('/', 7)]
        elif (url.find('/', 7) == -1):  #无资源路径且末尾无斜杠
            self.host = url[7:]
            self.obj = '/'
            self.host_protocol = url
        if (self.protocol(url) == 2 and url.find('/', 8) != -1):  #判断是https协议
            self.host = url[8:url.find('/', 8)]
            self.obj = url[url.find('/', 8):]
            self.host_protocol = url[:url.find('/', 8)]
        elif (url.find('/', 8) == -1):  #无资源路径且末尾无斜杠
            self.host = url[8:]
            self.obj = '/'
            self.host_protocol = url
        self.host_complete = url
        if (len(self.host) == 0):
            print(self.__class__.__name__, '::', sys._getframe().f_code.co_name, '解析url错误：', url)
            return False
        else:
            return True

    def sendRequestSocket(self):  #(未完成)发送socket请求，获得网页源码
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  #AF_INET指定使用IPv4协议，SOCK_STREAM指定使用面向流的TCP协议
        except socket.error as msg:
            print ('创建scoket失败，错误代码: ' + str(msg[0]) + '，错误信息 : ' + msg[1])
            sys.exit()
        s.connect((self.host, self.port))

        info = 'GET / HTTP/1.1\r\n'
        info += 'Host: ' + self.host + '\r\n'
        #info += 'User-Agent: Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36\r\n'
        #info += "Cookie: buvid3=3224324A-22FA-4973-8769-8F4A24222456110257infoc; UM_distinctid=16b0c3204fd204-0a2e4dc1ac6008-454c092b-1fa400-16b0c3204fe442; CURRENT_FNVAL=16; rpdid=|(umYlk~RYmR0J'ullmk~~~l|; LIVE_BUVID=AUTO9215592802831035; stardustvideo=1; im_notify_type_93025254=0; _uuid=EBC27393-073F-5DAC-74B1-C8A4202F4F8218329infoc; DedeUserID=93025254; DedeUserID__ckMd5=55326bd66e767225; SESSDATA=7bf97a6d%2C1570078647%2Cecc84991; bili_jct=192a11a62cb999f0a416581ccd08bcab; sid=6pymgz13; CURRENT_QUALITY=64; bp_t_offset_93025254=303331300778090737\r\n"
        info += 'Connection: close\r\n\r\n'
        print(info, self.obj)
        
        s.send(info.encode('utf-8'))
        buffer = []
        while True:
            #每次最多接收1k字节:
            d = s.recv(1024)
            if d:
                buffer.append(d)
            else:
                break
        data = b''.join(buffer)
        s.close()
        return data.decode('utf-8')

    def sendRequestRequests(self, url=''):  #使用requests模块发送请求，获得网页源码
        if (len(url) == 0):
            if (spider_enable_proxy):
                response = requests.get(self.host_complete, headers = self.headers, proxies = self.proxy)  #使用代理，添加：proxies = self.proxy
            else:
                response = requests.get(self.host_complete, headers = self.headers)
        else:
            if (spider_enable_proxy):
                response = requests.get(url, headers = self.headers, proxies = self.proxy)
            else:
                response = requests.get(url, headers = self.headers)

        if (response.status_code != 200):  #url为空或数据错误返回失败
            print(self.__class__.__name__, '::', sys._getframe().f_code.co_name, '网站反馈失败，状态码:', response.status_code)
            return False

        self.tree = html.fromstring(response.content)  #在解析xml格式时，将字符串转换为Element对象，方便xpath定位
        #print(response.content.decode('utf-8'))
        #with open('./temp.html', 'w') as f:
        #    f.write(response.content.decode('utf-8'))
        return True

    def analysisContent(self, i):  #解析获取网页内容
        if (self.sendRequestRequests() == True):  #获得网页源码成功
            lists[i][3] = 0  #解除占用状态
            lists.pop(i)
            '''if ():  #是否为资源页，获得想要的关键信息
                #获取信息等
                global spider_num
                #写入Excel解除list占用，并删除本条list减少内存占用
                if (self.os_name == 'posix'):
                    print('\033[1;32m --(' + spider_num.__str__() + ')-- \033[0m', 'len(lists) =', len(lists), lists[i])
                if (self.os_name == 'nt'):
                    print('--(' + spider_num.__str__() + ')--', 'len(lists) =', len(lists), lists[i])
                #self.wb.active.append(lists[i])  #写入Excel
                #self.wb.save(self.file_complete_path)
                #解除占用状态
                #删除处理完成的成员，节省内存
                if (spider_num > 0):  #计算爬取个数
                    spider_num = spider_num - 1
            elif ():  #判断真在爬取的网址类型是否为主页，从主页获得要爬取的URL
                #解除占用状态
                if (self.os_name == 'posix'):
                    print('\033[1;35m 检索主页： \033[0m', '\033[1;34m ' + lists[i][2] + ' \033[0m')
                if (self.os_name == 'nt'):
                    print('检索主页：', lists[i][2])
                #删除处理完成的成员，节省内存
                #获取改主页的其他资源页、其他信息
                #判断该资源未被爬取过
                #获取下一页地址，写入爬取队列
                #未找到下一页，不再往爬取队列写入新值return'''
    
    def homePage(self):  #判断该链接是否为主页类型的链接
        count = 0
        for li in self.home_list:
            if (self.obj.find(li) != -1):
                count = count + 1
        return count > 0
    
    def run(self, url):
        lists.append(['', '', url, ''])
        i = 0
        while spider_num or spider_num == -1:
            lists[i][3] = 1  #设置本条链接为占用状态
            if (self.analysis(lists[i][2]) == True):  #解析URL，获得主机名和资源路径
                self.analysisContent(i)  #解析获取网页内容
            else:
                return
            if (len(lists) == 0):
                print('遍历完成！')
                return
            #i = i + 1  #这里先不做处理，以后多线程调用可能会使用

def main(argv):
    global spider_num, spider_url, spider_enable_proxy
    try:  #读取命令行选项
        opts, args = getopt.getopt(argv, "hxn:u:", ["number=","url="])
    except getopt.GetoptError:
        print (sys.argv[0] + ' -n <number> -u <url>')
        sys.exit(2)
    for opt, arg in opts:
        if (opt == '-h'):
            print (sys.argv[0] + ' -n <number> -u <url>')
            sys.exit()
        elif (opt in ("-n", "--number")):
            spider_num = int(arg)
        elif (opt in ("-u", "--url")):
            spider_url = arg
        elif (opt == '-x'):
            spider_enable_proxy = True
    if (spider_num <= -1):
        print ('本次任务量不上限')
    else:
        print ('本次任务量：', spider_num, '条')
    print ('本次目标链接：', spider_url)

    splider = MySplider()
    splider.run(spider_url)  #运行爬虫

if (__name__=='__main__'):
    main(sys.argv[1:])
